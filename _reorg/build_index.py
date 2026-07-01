import os, re, csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def preview(p):
    ext=p.lower().rsplit('.',1)[-1]
    try:
        if ext=='docx':
            from docx import Document
            t=" ".join(x.text for x in Document(p).paragraphs if x.text.strip())
        elif ext=='pdf':
            import subprocess
            t=subprocess.run(['pdftotext','-l','1',p,'-'],capture_output=True,text=True,timeout=30).stdout
        elif ext in ('txt','eml','md','html','csv'):
            t=open(p,errors='ignore').read()
        else:
            return ""
        return re.sub(r'\s+',' ',t).strip()[:220]
    except Exception:
        return ""

def ddate(name):
    m=re.search(r'(\b\d{1,2})[._/](\d{1,2})[._/](\d{2,4})\b',name)
    if m:
        mo,da,yr=m.groups(); yr=('20'+yr) if len(yr)==2 else yr
        try: return f"{int(yr):04d}-{int(mo):02d}-{int(da):02d}"
        except: pass
    m=re.search(r'\b(20\d{2})\b',name); return m.group(1) if m else ""

# ---- curated overrides (filename -> [tags, direction, weight, what_it_proves]) ----
C={
"436-2026-00670_CDI_ChargeOfDiscrimination.pdf":["L1;L2;L3;L4","Supports","Direct","EEOC Charge of Discrimination vs Contender (#436-2026-00670) — operative ADEA charge."],
"436-2026-00698_BIRE_ChargeOfDiscrimination.pdf":["L1;L2;L3;L4","Supports","Direct","EEOC Charge of Discrimination vs Blackstream (#436-2026-00698)."],
"Blackstream - Lord - Statement of Position to EEOC 6.24.26.pdf":["L1;L3;L4;F3;F4;F7;F13;F26","Adverse","Direct","Respondent Position Statement (Fisher & Phillips, 6/24/26). Denies ADEA; frames transfer as lateral, warning for false statements; stray-remark & same-class defenses; falsely asserts Lord still employed."],
"EEOC_FisherPhillips Response.pdf":["L1;L3","Adverse","Direct","Respondent counsel (Fisher & Phillips) response."],
"EEOC_Rebuttal_Blackstream.CDI_6_30_26.docx":["L1;L2;L3;L4;L5;L7;L8","Supports","Direct","Lord's comprehensive rebuttal (original). Discloses 5/15/26 termination; rebuts each defense; alt. constructive-discharge argument."],
"EEOC_Rebuttal_Blackstream_Breck_Seiniger_review1_CDI_6_30_26.docx":["L1;L2;L3;L4;L5;L7;L8","Supports","Direct","Review edit of the rebuttal (Breck Seiniger pass)."],
"Formal Notice of Claims 11.14.2025.pdf":["L1;L2;L3;L4;L7","Supports","Direct","Beth's 11/14/25 Formal Notice of Claims + preservation demand to T. Phillips — anchors protected activity & preservation duty."],
"EEOC Investigator Summary Brief 5.25.26.docx":["L1;L2;L3;L4","Supports","Strong","Lord's summary brief to the EEOC investigator (5/25/26)."],
"EEOC Complaint Overview 5.21.26.docx":["L1;L2;L3;L4","Supports","Strong","Overview of Lord's complaint for the EEOC investigator (5/21/26)."],
"NLRB Statement 3.9.26.docx":["L5;L3;F13;F24","Supports","Direct","Section 7 statement: protected concerted activity (safety, payroll) + retaliation timeline."],
"CHG.10-CA-381639.First Amended Charge Against Employer - CA Case.pdf":["L5","Supports","Direct","NLRB First Amended Charge Against Employer (10-CA-381639)."],
"NLRB Charge.png":["L5","Supports","Direct","NLRB charge filing / confirmation."],
"NLRB Witness List 3.1.26.docx":["L5","Supports","Strong","Witness list for the NLRB charge."],
"Lord v CDI Audio Transcription of Disc Action.docx":["L3;F13","Supports","Direct","Transcript of the disciplinary (verbal warning) meeting — timing/basis of the write-up."],
"Lord v CDI Audio Transcription of Review.docx":["F15;L3","Supports","Direct","Transcript of the Dec 15 performance-review meeting ('excellent' review)."],
}
# exhibit filenames vary slightly; match by prefix key
EX=[
("EXHIBIT A","F1","Supports","Direct","CEO org chart showing Cape & Weldon reporting to Lord — rebuts 'Cape didn't report to Lord' (F1)."),
("EXHIBIT B","F1","Supports","Direct","HR (Aldridge) 3/13/25 memo confirming Lord's supervisory & budget authority (F1)."),
("EXHIBIT C","F4;F19;F25","Supports","Strong","Notes of 3/19 Mumma meeting revoking Lord's authority."),
("EXHIBIT D","F3;F19;F24","Supports","Strong","4/15/25 email to HR raising concerns re forced transfer/payroll (protected activity)."),
("EXHIBIT E","F4;F19","Supports","Corroborating","Office reassignment / removal-of-authority evidence."),
("EXHIBIT F","F21","Supports","Corroborating","Wrenn title change (6/25/25)."),
("EXHIBIT G","L7;F18","Supports","Strong","9/2/25 email lockout (password not reset) — spoliation."),
("EXHIBIT H","F7;F21","Supports","Direct","9/5/25 Brodsky email introducing Cape as Director of Marketing — younger employee assumed role."),
("EXHIBIT I","F14","Supports","Corroborating","Investment-news article re company financial issues (2/25/26)."),
("EXHIBIT J","F14;F19","Supports","Corroborating","Email to TP re reputational risk (2/25/26)."),
("EXHIBIT K","F19","Supports","Strong","90.3% non-response rate to Lord's work product (May–Nov 2025)."),
("EXHIBIT X","F1","Supports","Corroborating","Time-off approval showing Lord's authority over Cape."),
]

def heuristic(folder,name):
    n=name.lower()
    if 'position statement' in n or 'fisher' in n: return ("L1;L3;L4","Adverse","Direct")
    if 'rebuttal' in n: return ("L1;L2;L3;L4","Supports","Strong")
    if 'charge' in n and 'discrimination' in n: return ("L1","Supports","Direct")
    if 'nlrb' in n: return ("L5","Supports","Strong")
    if 'timeline' in n: return ("L1;L3;L4","Supports","Background")
    if 'preservation' in n: return ("L7","Supports","Strong")
    if 'notice of claims' in n or 'formal complant' in n or 'formal complaint' in n: return ("L1;L3;L4","Supports","Strong")
    if 'review' in n or 'performance' in n: return ("F15;L3","Supports","Strong")
    if 'transcription' in n or 'audio' in n: return ("L3","Supports","Strong")
    if 'lockout' in n: return ("L7;F18","Supports","Strong")
    if 'wrenn' in n or 'annette' in n: return ("F20;F21;F30","Supports","Corroborating")
    if 'cape' in n or 'promotion' in n or 'hayley' in n or 'haley' in n: return ("F7;F21","Supports","Corroborating")
    if 'gun' in n or 'weapon' in n or 'cam ' in n: return ("F25","Supports","Corroborating")
    if 'proposal' in n or '1099' in n or 'contractor' in n or 'cmo' in n or 'vpm' in n: return ("F5;F6","Supports","Strong")
    if 'handbook' in n: return ("F27","Adverse","Background")
    if 'reporting' in n or 'metrics' in n or 'sm ' in n: return ("F1;F2","Supports","Corroborating")
    if 'budget' in n: return ("F19","Supports","Corroborating")
    if 'org' in n or 'authority' in n: return ("F1","Supports","Strong")
    if folder.endswith('emails'): return ("","Supports","Corroborating")
    if folder.endswith('texts'): return ("","Supports","Corroborating")
    if folder.endswith('images-screenshots'): return ("","Supports","Corroborating")
    return ("","Supports","Background")

rows=[]
for folder,_,fns in os.walk('.'):
    top=folder.split('/')[1] if '/' in folder else folder
    if top not in ("00_CASE_SUMMARY","01_FILINGS","02_TIMELINES","03_EVIDENCE","04_EXHIBITS","05_CORRESPONDENCE","06_WORK_PRODUCT"): continue
    for name in sorted(fns):
        if name=='.DS_Store' or name.startswith('~$') or name.endswith('.md'): continue
        p=os.path.join(folder,name)
        ext=name.rsplit('.',1)[-1].lower() if '.' in name else ''
        cat=folder.lstrip('./')
        if name in C:
            tags,dr,wt,wip=C[name]; conf="curated"
        else:
            hit=None
            for pref,tg,d2,w2,wi in EX:
                if name.startswith(pref): hit=(tg,d2,w2,wi); break
            if hit:
                tags,dr,wt,wip=hit; conf="curated"
            else:
                tags,dr,wt=heuristic(cat,name); wip=""; conf="provisional — review"
        rows.append([ "", name, cat, ddate(name), ext.upper(), tags, dr, wt, wip, conf, preview(p) ])

# sort: filings, timelines, exhibits, evidence, work product, summary
order={"01_FILINGS/EEOC":0,"01_FILINGS/NLRB":1,"02_TIMELINES":2,"04_EXHIBITS":3,"03_EVIDENCE/emails":4,"03_EVIDENCE/texts":5,"03_EVIDENCE/documents":6,"03_EVIDENCE/audio-transcripts":7,"03_EVIDENCE/images-screenshots":8,"06_WORK_PRODUCT":9,"00_CASE_SUMMARY":10,"05_CORRESPONDENCE":11}
rows.sort(key=lambda r:(order.get(r[2],99), r[3] or "zzz", r[1]))

wb=Workbook(); ws=wb.active; ws.title="Document Index"
hdr=["Exhibit #","File","Folder / Category","Date","Type","Issue Tags","Direction","Weight","What It Proves","Confidence","Content Preview"]
ws.append(hdr)
navy=PatternFill("solid",fgColor="1F3864"); white=Font(bold=True,color="FFFFFF",name="Arial",size=10)
thin=Side(style="thin",color="D9D9D9"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
for c in ws[1]:
    c.fill=navy; c.font=white; c.alignment=Alignment(vertical="center",horizontal="left",wrap_text=True); c.border=border
supfill=PatternFill("solid",fgColor="E2EFDA"); advfill=PatternFill("solid",fgColor="FCE4D6"); mixfill=PatternFill("solid",fgColor="FFF2CC")
prov=Font(name="Arial",size=9,italic=True,color="808080"); normal=Font(name="Arial",size=9)
for r in rows:
    ws.append(r)
    row=ws[ws.max_row]
    for c in row:
        c.alignment=Alignment(vertical="top",wrap_text=True); c.border=border; c.font=normal
    dcell=row[6]  # Direction
    if r[6]=="Supports": dcell.fill=supfill
    elif r[6]=="Adverse": dcell.fill=advfill
    elif r[6]=="Mixed": dcell.fill=mixfill
    if r[9].startswith("provisional"): row[9].font=prov
widths=[9,42,26,11,7,16,11,13,50,17,60]
from openpyxl.utils import get_column_letter
for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:K{ws.max_row}"
os.makedirs("00_CASE_SUMMARY",exist_ok=True)
wb.save("00_CASE_SUMMARY/DOCUMENT_INDEX.xlsx")
cur=sum(1 for r in rows if r[9]=="curated"); print(f"rows: {len(rows)} | curated: {cur} | provisional: {len(rows)-cur}")
